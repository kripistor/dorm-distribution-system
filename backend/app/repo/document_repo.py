from datetime import timezone

from openpyxl.styles import Font, Alignment
from openpyxl.workbook import Workbook
from sqlalchemy import select, join

from app.models import UserProfile, PersonAttachedRoom, Room
from app.repo.dormitory_repo import DormitoryRepo
from app.repo.repo import SQLAlchemyRepo


class DocumentRepo(SQLAlchemyRepo):
    async def generate_report(self):
        workbook = Workbook()
        workbook.remove(workbook.active)

        dormitory_repo = DormitoryRepo(self.session)
        dormitories = await dormitory_repo.get_all()

        bold_font = Font(bold=True)
        left_alignment = Alignment(horizontal="left")
        center_alignment = Alignment(horizontal="center")

        for dormitory in dormitories:
            dormitory_statistics = await dormitory_repo.get_dormitory_statistics(
                dormitory.id
            )

            for floor in dormitory_statistics.floors:
                floor_sheet = workbook.create_sheet(title=f"Этаж № {floor.floor_name}")

                dormitory_info = [
                    "Название общажития",
                    "Адрес общажития",
                    "Всего мест",
                    "Занято мест",
                ]
                floor_sheet.append(dormitory_info)
                for row in floor_sheet["A1":"D1"]:
                    for cell in row:
                        cell.font = bold_font
                        cell.alignment = left_alignment
                floor_sheet.append(
                    [
                        dormitory.name,
                        dormitory.address,
                        dormitory_statistics.total_space,
                        dormitory_statistics.occupied_space,
                    ]
                )

                floor_info = ["Этаж №", "Всего мест", "Занято мест", "Свободно комнат"]
                floor_sheet.append(floor_info)
                for row in floor_sheet["A3":"D3"]:
                    for cell in row:
                        cell.font = bold_font
                        cell.alignment = left_alignment
                floor_sheet.append(
                    [
                        floor.floor_name,
                        floor.total_space,
                        floor.occupied_space,
                        floor.rooms_free,
                    ]
                )

                rooms = await self.session.execute(
                    select(Room).where(Room.floor_id == floor.floor_id)
                )
                rooms = rooms.scalars().all()

                for room in rooms:
                    room_info = ["Комната №", "Вместимость команты"]
                    floor_sheet.append(room_info)
                    for row in floor_sheet[
                        f"A{floor_sheet.max_row}":f"B{floor_sheet.max_row}"
                    ]:
                        for cell in row:
                            cell.font = bold_font
                            cell.alignment = left_alignment
                    floor_sheet.append([room.id, room.capacity])

                    users = (
                        (
                            await self.session.execute(
                                select(UserProfile)
                                .select_from(
                                    join(
                                        UserProfile,
                                        PersonAttachedRoom,
                                        UserProfile.user_id
                                        == PersonAttachedRoom.user_id,
                                    )
                                )
                                .where(PersonAttachedRoom.room_id == room.id)
                            )
                        )
                        .scalars()
                        .all()
                    )

                    for user in users:
                        birth_date = user.birth_date.astimezone(timezone.utc).replace(
                            tzinfo=None
                        )
                        transfer_date = user.transfer_date.astimezone(
                            timezone.utc
                        ).replace(tzinfo=None)
                        user_info = [
                            "Айди Студента",
                            "Номер студенческого ",
                            "Имя",
                            "Дата рождения",
                            "Телефон",
                            "Номер приказа о зачислении",
                            "Номер приказа о заселении в общажитие",
                            "Дата заселения в общежитие",
                            "Место рождения",
                            "Адрес проживания",
                            "Инвалидность",
                            "Пол",
                        ]
                        floor_sheet.append(user_info)
                        for row in floor_sheet[
                            f"A{floor_sheet.max_row}":f"L{floor_sheet.max_row}"
                        ]:
                            for cell in row:
                                cell.font = bold_font
                                cell.alignment = left_alignment
                        floor_sheet.append(
                            [
                                str(user.user_id),
                                user.card_number,
                                user.name,
                                birth_date,
                                user.phone,
                                user.grant_order_number,
                                user.transfer_order_number,
                                transfer_date,
                                user.birth_place,
                                user.address,
                                user.concession,
                                user.gender,
                            ]
                        )

        filename = "report.xlsx"
        workbook.save(filename)

        return filename
